from fastapi import FastAPI, APIRouter, HTTPException, Query, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
import asyncio
from concurrent.futures import ThreadPoolExecutor
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
from enum import Enum
import bcrypt
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
import json
import openpyxl
import openpyxl.styles
from io import BytesIO
from fastapi.responses import Response

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Security configurations
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-this-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI(title="Kawale Cranes API", description="Data entry system for Kawale Cranes orders with authentication")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# User Roles Enum
class UserRole(str, Enum):
    SUPER_ADMIN = "super_admin"
    ADMIN = "admin"
    DATA_ENTRY = "data_entry"

# Define Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    full_name: str
    role: UserRole
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None
    last_login: Optional[datetime] = None

class UserCreate(BaseModel):
    email: EmailStr
    full_name: str
    password: str
    role: UserRole

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    role: Optional[UserRole] = None
    is_active: Optional[bool] = None

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    action: str  # CREATE, UPDATE, DELETE, LOGIN, LOGOUT
    resource_type: str  # USER, ORDER
    resource_id: Optional[str] = None
    old_data: Optional[Dict[str, Any]] = None
    new_data: Optional[Dict[str, Any]] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None

class CraneOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    added_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    ip_address: Optional[str] = None
    unique_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    customer_name: str
    phone: str
    order_type: str  # "cash" or "company"
    
    # Audit fields
    created_by: Optional[str] = None
    updated_by: Optional[str] = None
    updated_at: Optional[datetime] = None
    
    # Incentive fields (admin only)
    incentive_amount: Optional[float] = None
    incentive_reason: Optional[str] = None
    incentive_added_by: Optional[str] = None
    incentive_added_at: Optional[datetime] = None
    
    # Cash order fields
    cash_trip_from: Optional[str] = None
    cash_trip_to: Optional[str] = None
    care_off: Optional[str] = None
    care_off_amount: Optional[float] = None
    cash_vehicle_details: Optional[str] = None
    cash_driver_details: Optional[str] = None
    cash_vehicle_name: Optional[str] = None
    cash_vehicle_number: Optional[str] = None
    cash_service_type: Optional[str] = None
    amount_received: Optional[float] = None
    advance_amount: Optional[float] = None
    cash_kms_travelled: Optional[float] = None
    cash_toll: Optional[float] = None
    diesel: Optional[str] = None
    cash_diesel: Optional[float] = None
    cash_diesel_refill_location: Optional[str] = None
    cash_driver_name: Optional[str] = None
    cash_towing_vehicle: Optional[str] = None
    
    # Company order fields
    name_of_firm: Optional[str] = None
    company_name: Optional[str] = None  # Mandatory via endpoint validation
    case_id_file_number: Optional[str] = None
    company_vehicle_name: Optional[str] = None
    company_vehicle_number: Optional[str] = None
    company_service_type: Optional[str] = None  # Mandatory via endpoint validation
    company_vehicle_details: Optional[str] = None
    company_driver_details: Optional[str] = None  # Mandatory via endpoint validation (Driver)
    company_trip_from: Optional[str] = None
    company_trip_to: Optional[str] = None
    reach_time: Optional[datetime] = None
    drop_time: Optional[datetime] = None
    company_kms_travelled: Optional[float] = None
    company_toll: Optional[float] = None
    diesel_name: Optional[str] = None
    company_diesel: Optional[float] = None
    company_diesel_refill_location: Optional[str] = None
    company_driver_name: Optional[str] = None
    company_towing_vehicle: Optional[str] = None  # Mandatory via endpoint validation (Towing Vehicle)

class CraneOrderCreate(BaseModel):
    customer_name: str
    phone: str
    order_type: str
    date_time: Optional[datetime] = None
    
    # Cash order fields
    cash_trip_from: Optional[str] = None
    cash_trip_to: Optional[str] = None
    care_off: Optional[str] = None
    care_off_amount: Optional[float] = None
    cash_vehicle_details: Optional[str] = None
    cash_driver_details: Optional[str] = None
    cash_vehicle_name: Optional[str] = None
    cash_vehicle_number: Optional[str] = None
    cash_service_type: Optional[str] = None
    amount_received: Optional[float] = None
    advance_amount: Optional[float] = None
    cash_kms_travelled: Optional[float] = None
    cash_toll: Optional[float] = None
    diesel: Optional[str] = None
    cash_diesel: Optional[float] = None
    cash_diesel_refill_location: Optional[str] = None
    cash_driver_name: Optional[str] = None
    cash_towing_vehicle: Optional[str] = None
    
    # Company order fields
    name_of_firm: Optional[str] = None
    company_name: Optional[str] = None  # Mandatory via endpoint validation
    case_id_file_number: Optional[str] = None
    company_vehicle_name: Optional[str] = None
    company_vehicle_number: Optional[str] = None
    company_service_type: Optional[str] = None  # Mandatory via endpoint validation
    company_vehicle_details: Optional[str] = None
    company_driver_details: Optional[str] = None  # Mandatory via endpoint validation (Driver)
    company_trip_from: Optional[str] = None
    company_trip_to: Optional[str] = None
    reach_time: Optional[datetime] = None
    drop_time: Optional[datetime] = None
    company_kms_travelled: Optional[float] = None
    company_toll: Optional[float] = None
    diesel_name: Optional[str] = None
    company_diesel: Optional[float] = None
    company_diesel_refill_location: Optional[str] = None
    company_driver_name: Optional[str] = None
    company_towing_vehicle: Optional[str] = None  # Mandatory via endpoint validation (Towing Vehicle)
    
    # Incentive fields (admin only)
    incentive_amount: Optional[float] = None
    incentive_reason: Optional[str] = None
    incentive_added_by: Optional[str] = None
    incentive_added_at: Optional[datetime] = None

class CraneOrderUpdate(BaseModel):
    customer_name: Optional[str] = None
    phone: Optional[str] = None
    order_type: Optional[str] = None
    date_time: Optional[datetime] = None
    
    # Cash order fields
    cash_trip_from: Optional[str] = None
    cash_trip_to: Optional[str] = None
    care_off: Optional[str] = None
    care_off_amount: Optional[float] = None
    cash_vehicle_details: Optional[str] = None
    cash_driver_details: Optional[str] = None
    cash_vehicle_name: Optional[str] = None
    cash_vehicle_number: Optional[str] = None
    cash_service_type: Optional[str] = None
    amount_received: Optional[float] = None
    advance_amount: Optional[float] = None
    cash_kms_travelled: Optional[float] = None
    cash_toll: Optional[float] = None
    diesel: Optional[str] = None
    cash_diesel: Optional[float] = None
    cash_diesel_refill_location: Optional[str] = None
    cash_driver_name: Optional[str] = None
    cash_towing_vehicle: Optional[str] = None
    
    # Company order fields
    name_of_firm: Optional[str] = None
    company_name: Optional[str] = None  # Mandatory in main model but optional for updates
    case_id_file_number: Optional[str] = None
    company_vehicle_name: Optional[str] = None
    company_vehicle_number: Optional[str] = None
    company_service_type: Optional[str] = None  # Mandatory in main model but optional for updates
    company_vehicle_details: Optional[str] = None
    company_driver_details: Optional[str] = None  # Mandatory in main model but optional for updates
    company_trip_from: Optional[str] = None
    company_trip_to: Optional[str] = None
    reach_time: Optional[datetime] = None
    drop_time: Optional[datetime] = None
    company_kms_travelled: Optional[float] = None
    company_toll: Optional[float] = None
    diesel_name: Optional[str] = None
    company_diesel: Optional[float] = None
    company_diesel_refill_location: Optional[str] = None
    company_driver_name: Optional[str] = None
    company_towing_vehicle: Optional[str] = None  # Mandatory in main model but optional for updates
    
    # Incentive fields (admin only)
    incentive_amount: Optional[float] = None
    incentive_reason: Optional[str] = None
    incentive_added_by: Optional[str] = None
    incentive_added_at: Optional[datetime] = None

class ServiceRate(BaseModel):
    model_config = ConfigDict(extra='forbid')
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name_of_firm: str
    company_name: str
    service_type: str
    base_rate: float  # Rate for base distance (up to 40km)
    base_distance_km: float = Field(default=40.0)  # Base distance in km
    rate_per_km_beyond: float  # Rate per km beyond base distance
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CalculatedOrderFinancials(BaseModel):
    """Calculated financial data for an order"""
    base_revenue: float = 0.0  # Revenue from rate calculation
    incentive_amount: float = 0.0  # Incentive paid
    total_revenue: float = 0.0  # base_revenue + incentive_amount
    calculation_details: Optional[str] = None  # Details of how calculation was done

# Helper functions for MongoDB serialization


class ImportHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    filename: str
    imported_by: str
    imported_by_email: str
    imported_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    total_records: int
    success_count: int
    error_count: int
    cash_orders: int = 0
    company_orders: int = 0


class DriverSalary(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    driver_name: str
    month: int  # 1-12
    year: int
    base_salary: float
    total_incentives: float = 0.0
    deductions: float = 0.0
    notes: Optional[str] = None
    added_by: str
    added_by_email: str
    added_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class DriverDefaultSalary(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    driver_name: str
    default_salary: float = 15000.0
    updated_by: str
    updated_by_email: str
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

    updated_by: Optional[str] = None
    updated_at: Optional[datetime] = None

    sample_data: List[Dict[str, Any]] = []  # First 5 records as preview

def prepare_for_mongo(data):
    """Convert datetime objects to ISO strings for MongoDB storage"""
    doc = dict(data)
    for key, value in doc.items():
        if isinstance(value, datetime):
            doc[key] = value.isoformat()
    return doc

def parse_from_mongo(item):
    """Convert ISO strings back to datetime objects from MongoDB and handle None values"""
    if not item:
        return item
    
    # Convert datetime fields
    datetime_fields = ['added_time', 'date_time', 'reach_time', 'drop_time', 'created_at', 'updated_at', 'last_login', 'timestamp', 'incentive_added_at']
    for field in datetime_fields:
        if field in item and isinstance(item[field], str):
            try:
                item[field] = datetime.fromisoformat(item[field])
            except ValueError:
                continue
    
    # Handle None values in string fields that should be Optional[str] but not None for Pydantic validation
    # Convert None to empty string for optional string fields to avoid Pydantic validation errors
    optional_string_fields = [
        'cash_trip_from', 'cash_trip_to', 'care_off', 'cash_vehicle_details', 'cash_driver_details',
        'cash_vehicle_name', 'cash_vehicle_number', 'cash_service_type', 'diesel', 'cash_diesel_refill_location',
        'cash_driver_name', 'cash_towing_vehicle', 'name_of_firm', 'company_name', 'case_id_file_number',
        'company_vehicle_name', 'company_vehicle_number', 'company_service_type', 'company_vehicle_details',
        'company_driver_details', 'company_trip_from', 'company_trip_to', 'diesel_name',
        'company_diesel_refill_location', 'company_driver_name', 'company_towing_vehicle', 'incentive_reason'
    ]
    
    for field in optional_string_fields:
        if field in item and item[field] is None:
            item[field] = ""  # Convert None to empty string
    
    return item

# Authentication utility functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_user_by_email(email: str):
    user = await db.users.find_one({"email": email}, {"_id": 0})
    return parse_from_mongo(user) if user else None

async def authenticate_user(email: str, password: str):
    user = await get_user_by_email(email)
    if not user:
        return False
    
    # Get user with password for authentication
    user_with_password = await db.users.find_one({"email": email})
    if not user_with_password or not verify_password(password, user_with_password["hashed_password"]):
        return False
    return user

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await get_user_by_email(email)
    if user is None:
        raise credentials_exception
    
    if not user["is_active"]:
        raise HTTPException(status_code=400, detail="Inactive user")
    
    return user

def require_role(allowed_roles: List[UserRole]):
    def role_checker(current_user: dict = Depends(get_current_user)):
        if current_user["role"] not in [role.value for role in allowed_roles]:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not enough permissions"
            )
        return current_user
    return role_checker

async def log_audit(user_id: str, user_email: str, action: str, resource_type: str, 
                   resource_id: str = None, old_data: Dict = None, new_data: Dict = None,
                   ip_address: str = None, user_agent: str = None):
    """Log audit trail"""
    audit_log = AuditLog(
        user_id=user_id,
        user_email=user_email,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        old_data=old_data,
        new_data=new_data,
        ip_address=ip_address,
        user_agent=user_agent
    )
    
    doc = prepare_for_mongo(audit_log.model_dump())
    await db.audit_logs.insert_one(doc)

# API Endpoints
@api_router.get("/")
async def root():
    return {"message": "Kawale Cranes Data Entry System with Authentication"}

# Authentication endpoints
@api_router.post("/auth/register", response_model=User)
async def register_user(user_data: UserCreate, current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))):
    """Register a new user (Admin and Super Admin only)"""
    # Check if user already exists
    existing_user = await get_user_by_email(user_data.email)
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Hash password
    hashed_password = get_password_hash(user_data.password)
    
    # Create user
    user = User(
        email=user_data.email,
        full_name=user_data.full_name,
        role=user_data.role,
        created_by=current_user["id"]
    )
    
    # Store user with hashed password
    user_doc = prepare_for_mongo(user.model_dump())
    user_doc["hashed_password"] = hashed_password
    
    await db.users.insert_one(user_doc)
    
    # Log audit
    await log_audit(
        user_id=current_user["id"],
        user_email=current_user["email"],
        action="CREATE",
        resource_type="USER",
        resource_id=user.id,
        new_data={"email": user.email, "full_name": user.full_name, "role": user.role}
    )
    
    return user

@api_router.post("/auth/login", response_model=Token)
async def login(user_credentials: UserLogin):
    """Authenticate user and return token"""
    user = await authenticate_user(user_credentials.email, user_credentials.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Update last login
    await db.users.update_one(
        {"email": user["email"]},
        {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Create access token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["email"]}, expires_delta=access_token_expires
    )
    
    # Log audit
    await log_audit(
        user_id=user["id"],
        user_email=user["email"],
        action="LOGIN",
        resource_type="USER",
        resource_id=user["id"]
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": User(**user)
    }

@api_router.get("/auth/me", response_model=User)
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    """Get current user information"""
    return User(**current_user)

@api_router.post("/auth/logout")
async def logout(current_user: dict = Depends(get_current_user)):
    """Logout user (mainly for audit trail)"""
    # Log audit
    await log_audit(
        user_id=current_user["id"],
        user_email=current_user["email"],
        action="LOGOUT",
        resource_type="USER",
        resource_id=current_user["id"]
    )
    
    return {"message": "Logged out successfully"}


@api_router.put("/auth/change-password")
async def change_password(
    password_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Change current user's password"""
    try:
        # Validate input
        current_password = password_data.get("current_password")
        new_password = password_data.get("new_password")
        
        if not current_password or not new_password:
            raise HTTPException(status_code=400, detail="Current password and new password are required")
        
        if len(new_password) < 6:
            raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
        
        # Get user from database
        user = await db.users.find_one({"id": current_user["id"]})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Verify current password
        if not bcrypt.checkpw(current_password.encode('utf-8'), user["hashed_password"].encode('utf-8')):
            raise HTTPException(status_code=400, detail="Current password is incorrect")
        
        # Check if new password is same as current
        if current_password == new_password:
            raise HTTPException(status_code=400, detail="New password must be different from current password")
        
        # Hash new password
        hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Update password
        result = await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": {"hashed_password": hashed_password}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="UPDATE",
            resource_type="USER",
            resource_id=current_user["id"],
            old_data={"password": "***"},
            new_data={"password": "***", "changed_by": "self"}
        )
        
        return {"message": "Password changed successfully"}
    
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Error changing password: {str(e)}")


# User management endpoints
@api_router.get("/users", response_model=List[User])
async def get_users(
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN])),
    limit: int = Query(100, ge=1, le=1000)
):
    """Get all users (Admin and Super Admin only)"""
    users = await db.users.find({}, {"_id": 0, "hashed_password": 0}).limit(limit).to_list(limit)
    return [User(**parse_from_mongo(user)) for user in users]

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(
    user_id: str,
    update_data: UserUpdate,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Update user (Admin and Super Admin only)"""
    # Get existing user
    existing_user = await db.users.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prepare update
    update_dict = update_data.model_dump(exclude_unset=True, exclude_none=True)
    if update_dict:
        prepared_update = prepare_for_mongo(update_dict)
        
        # Update user
        result = await db.users.update_one(
            {"id": user_id},
            {"$set": prepared_update}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="UPDATE",
            resource_type="USER",
            resource_id=user_id,
            old_data=existing_user,
            new_data=update_dict
        )
    
    # Return updated user
    updated_user = await db.users.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
    return User(**parse_from_mongo(updated_user))

@api_router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Delete user (Super Admin only)"""
    try:
        # Prevent self-deletion
        if user_id == current_user["id"]:
            raise HTTPException(status_code=400, detail="Cannot delete your own account")
        
        # Get user for audit log
        existing_user = await db.users.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
        if not existing_user:
            raise HTTPException(status_code=404, detail="User not found")
        
        result = await db.users.delete_one({"id": user_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="DELETE",
            resource_type="USER",
            resource_id=user_id,
            old_data=existing_user
        )
        
        return {"message": "User deleted successfully"}
    
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Error deleting user: {str(e)}")


@api_router.put("/users/{user_id}/reset-password")
async def reset_user_password(
    user_id: str,
    password_data: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Reset user password (Super Admin only)"""
    try:
        # Validate password data
        new_password = password_data.get("new_password")
        if not new_password or len(new_password) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        
        # Prevent resetting own password through this endpoint
        if user_id == current_user["id"]:
            raise HTTPException(status_code=400, detail="Use change password endpoint for your own password")
        
        # Check if user exists
        existing_user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not existing_user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Hash new password
        hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Update password
        result = await db.users.update_one(
            {"id": user_id},
            {"$set": {"hashed_password": hashed_password}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Log audit (without storing the actual password)
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="UPDATE",
            resource_type="USER",
            resource_id=user_id,
            old_data={"password": "***"},
            new_data={"password": "***", "reset_by": current_user["email"]}
        )
        
        return {"message": f"Password reset successfully for user {existing_user.get('email')}"}
    
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Error resetting password: {str(e)}")

# Order endpoints (with authentication and audit)
@api_router.post("/orders", response_model=CraneOrder)
async def create_order(
    input: CraneOrderCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new crane order"""
    order_dict = input.model_dump(exclude_unset=True)
    
    # Set default datetime if not provided
    if not order_dict.get('date_time'):
        order_dict['date_time'] = datetime.now(timezone.utc)
    
    # Add audit fields
    order_dict['created_by'] = current_user["id"]
    
    order_obj = CraneOrder(**order_dict)
    
    # Validate mandatory fields for company orders
    if order_obj.order_type == "company":
        missing_fields = []
        
        if not order_obj.company_name or order_obj.company_name.strip() == "":
            missing_fields.append("Company Name")
        
        if not order_obj.company_service_type or order_obj.company_service_type.strip() == "":
            missing_fields.append("Service Type")
        
        if not order_obj.company_driver_details or order_obj.company_driver_details.strip() == "":
            missing_fields.append("Driver")
        
        if not order_obj.company_towing_vehicle or order_obj.company_towing_vehicle.strip() == "":
            missing_fields.append("Towing Vehicle")
        
        if missing_fields:
            field_list = ", ".join(missing_fields)
            raise HTTPException(
                status_code=422, 
                detail=f"The following fields are required for company orders: {field_list}"
            )
    
    # Convert to dict and serialize datetime fields for MongoDB
    doc = prepare_for_mongo(order_obj.model_dump())
    
    try:
        result = await db.crane_orders.insert_one(doc)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="CREATE",
            resource_type="ORDER",
            resource_id=order_obj.id,
            new_data=order_dict
        )
        
        return order_obj
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating order: {str(e)}")

@api_router.get("/orders", response_model=List[CraneOrder])
async def get_orders(
    current_user: dict = Depends(get_current_user),
    order_type: Optional[str] = Query(None, description="Filter by order type (cash/company)"),
    customer_name: Optional[str] = Query(None, description="Filter by customer name"),
    phone: Optional[str] = Query(None, description="Filter by phone number"),
    limit: int = Query(100, ge=1, le=1000, description="Number of orders to return"),
    skip: int = Query(0, ge=0, description="Number of orders to skip")
):
    """Get all crane orders"""
    query = {}
    
    # Build query filters
    if order_type:
        query["order_type"] = order_type
    if customer_name:
        query["customer_name"] = {"$regex": customer_name, "$options": "i"}
    if phone:
        query["phone"] = {"$regex": phone, "$options": "i"}
    
    try:
        # Exclude MongoDB's _id field from results
        orders = await db.crane_orders.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
        
        # Parse datetime fields from MongoDB
        parsed_orders = [parse_from_mongo(order) for order in orders]
        
        return parsed_orders
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching orders: {str(e)}")

@api_router.get("/orders/{order_id}", response_model=CraneOrder)
async def get_order(
    order_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific crane order by ID"""
    try:
        order = await db.crane_orders.find_one({"id": order_id}, {"_id": 0})
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        return parse_from_mongo(order)
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Error fetching order: {str(e)}")

@api_router.put("/orders/{order_id}", response_model=CraneOrder)
async def update_order(
    order_id: str,
    update_data: CraneOrderUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a specific crane order"""
    try:
        # Check if order exists
        existing_order = await db.crane_orders.find_one({"id": order_id}, {"_id": 0})
        if not existing_order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Prepare update data
        update_dict = update_data.model_dump(exclude_unset=True, exclude_none=True)
        
        if update_dict:
            # Create a combined dict of existing order + updates for validation
            combined_data = {**existing_order, **update_dict}
            
            # Validate mandatory fields for company orders
            if combined_data.get("order_type") == "company":
                missing_fields = []
                
                company_name = combined_data.get("company_name", "")
                if not company_name or (isinstance(company_name, str) and company_name.strip() == ""):
                    missing_fields.append("Company Name")
                
                company_service_type = combined_data.get("company_service_type", "")
                if not company_service_type or (isinstance(company_service_type, str) and company_service_type.strip() == ""):
                    missing_fields.append("Service Type")
                
                company_driver_details = combined_data.get("company_driver_details", "")
                if not company_driver_details or (isinstance(company_driver_details, str) and company_driver_details.strip() == ""):
                    missing_fields.append("Driver")
                
                company_towing_vehicle = combined_data.get("company_towing_vehicle", "")
                if not company_towing_vehicle or (isinstance(company_towing_vehicle, str) and company_towing_vehicle.strip() == ""):
                    missing_fields.append("Towing Vehicle")
                
                if missing_fields:
                    field_list = ", ".join(missing_fields)
                    raise HTTPException(
                        status_code=422, 
                        detail=f"The following fields are required for company orders: {field_list}"
                    )
            
            # Add audit fields
            update_dict['updated_by'] = current_user["id"]
            update_dict['updated_at'] = datetime.now(timezone.utc)
            
            # Serialize datetime fields
            prepared_update = prepare_for_mongo(update_dict)
            
            # Update the order
            result = await db.crane_orders.update_one(
                {"id": order_id},
                {"$set": prepared_update}
            )
            
            if result.matched_count == 0:
                raise HTTPException(status_code=404, detail="Order not found")
            
            # Log audit
            await log_audit(
                user_id=current_user["id"],
                user_email=current_user["email"],
                action="UPDATE",
                resource_type="ORDER",
                resource_id=order_id,
                old_data=existing_order,
                new_data=update_dict
            )
        
        # Fetch and return updated order
        updated_order = await db.crane_orders.find_one({"id": order_id}, {"_id": 0})
        return parse_from_mongo(updated_order)
    
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Error updating order: {str(e)}")

@api_router.delete("/orders/{order_id}")
async def delete_order(
    order_id: str,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Delete a specific crane order (Admin and Super Admin only)"""
    try:
        # Get order for audit log
        existing_order = await db.crane_orders.find_one({"id": order_id}, {"_id": 0})
        if not existing_order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        result = await db.crane_orders.delete_one({"id": order_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="DELETE",
            resource_type="ORDER",
            resource_id=order_id,
            old_data=existing_order
        )
        
        return {"message": "Order deleted successfully"}
    
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Error deleting order: {str(e)}")


@api_router.delete("/orders/delete-all")
async def delete_all_orders(
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Delete all crane orders (Super Admin only - DANGEROUS!)"""
    try:
        # Count orders before deletion
        count = await db.crane_orders.count_documents({})
        
        if count == 0:
            raise HTTPException(status_code=404, detail="No orders found to delete")
        
        # Delete all orders
        result = await db.crane_orders.delete_many({})
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="DELETE_ALL",
            resource_type="ORDER",
            new_data={"deleted_count": result.deleted_count}
        )
        
        return {
            "message": f"Successfully deleted all {result.deleted_count} orders",
            "deleted_count": result.deleted_count
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting all orders: {str(e)}")


@api_router.get("/orders/stats/summary")
async def get_orders_summary(current_user: dict = Depends(get_current_user)):
    """Get summary statistics of orders"""
    try:
        pipeline = [
            {
                "$group": {
                    "_id": "$order_type",
                    "count": {"$sum": 1},
                    "total_amount": {
                        "$sum": {
                            "$cond": {
                                "if": {"$eq": ["$order_type", "cash"]},
                                "then": "$amount_received",
                                "else": 0
                            }
                        }
                    }
                }
            }
        ]
        
        stats = await db.crane_orders.aggregate(pipeline).to_list(None)
        
        # Calculate total orders
        total_orders = await db.crane_orders.count_documents({})
        
        return {
            "total_orders": total_orders,
            "by_type": stats
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching summary: {str(e)}")

# Audit endpoints
@api_router.get("/audit-logs", response_model=List[AuditLog])
async def get_audit_logs(
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN])),
    resource_type: Optional[str] = Query(None, description="Filter by resource type (USER/ORDER)"),
    action: Optional[str] = Query(None, description="Filter by action (CREATE/UPDATE/DELETE/LOGIN/LOGOUT)"),
    user_email: Optional[str] = Query(None, description="Filter by user email"),
    limit: int = Query(100, ge=1, le=1000, description="Number of logs to return"),
    skip: int = Query(0, ge=0, description="Number of logs to skip")
):
    """Get audit logs (Admin and Super Admin only)"""
    query = {}
    
    if resource_type:
        query["resource_type"] = resource_type
    if action:
        query["action"] = action
    if user_email:
        query["user_email"] = {"$regex": user_email, "$options": "i"}
    
    try:
        logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
        return [AuditLog(**parse_from_mongo(log)) for log in logs]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching audit logs: {str(e)}")

# Export endpoints
@api_router.get("/export/excel")
async def export_orders_excel(
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN])),
    order_type: Optional[str] = Query(None),
    customer_name: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    limit: int = Query(1000, ge=1, le=5000)
):
    """Export orders to Excel (Admin and Super Admin only)"""
    try:
        # Build query
        query = {}
        if order_type:
            query["order_type"] = order_type
        if customer_name:
            query["customer_name"] = {"$regex": customer_name, "$options": "i"}
        if phone:
            query["phone"] = {"$regex": phone, "$options": "i"}
        
        # Get orders
        orders = await db.crane_orders.find(query, {"_id": 0}).limit(limit).to_list(limit)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kawale Cranes Orders"
        
        # Headers
        headers = [
            "Order ID", "Date/Time", "Customer Name", "Phone", "Order Type",
            "Trip From", "Trip To", "Vehicle", "Driver", "Service Type",
            "Amount", "Advance", "KMs", "Toll", "Diesel", "Status"
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, order in enumerate(orders, 2):
            parsed_order = parse_from_mongo(order)
            
            ws.cell(row, 1, parsed_order.get("unique_id", "")[:8])
            ws.cell(row, 2, parsed_order.get("date_time", "").strftime("%Y-%m-%d %H:%M") if parsed_order.get("date_time") else "")
            ws.cell(row, 3, parsed_order.get("customer_name", ""))
            ws.cell(row, 4, parsed_order.get("phone", ""))
            ws.cell(row, 5, parsed_order.get("order_type", "").title())
            
            if parsed_order.get("order_type") == "cash":
                ws.cell(row, 6, parsed_order.get("cash_trip_from", ""))
                ws.cell(row, 7, parsed_order.get("cash_trip_to", ""))
                ws.cell(row, 8, parsed_order.get("cash_vehicle_name", ""))
                ws.cell(row, 9, parsed_order.get("cash_driver_name", ""))
                ws.cell(row, 10, parsed_order.get("cash_service_type", ""))
                ws.cell(row, 11, parsed_order.get("amount_received", 0))
                ws.cell(row, 12, parsed_order.get("advance_amount", 0))
                ws.cell(row, 13, parsed_order.get("cash_kms_travelled", 0))
                ws.cell(row, 14, parsed_order.get("cash_toll", 0))
                ws.cell(row, 15, parsed_order.get("cash_diesel", 0))
            else:
                ws.cell(row, 6, parsed_order.get("company_trip_from", ""))
                ws.cell(row, 7, parsed_order.get("company_trip_to", ""))
                ws.cell(row, 8, parsed_order.get("company_vehicle_name", ""))
                ws.cell(row, 9, parsed_order.get("company_driver_name", ""))
                ws.cell(row, 10, parsed_order.get("company_service_type", ""))
                ws.cell(row, 11, "")  # No amount for company
                ws.cell(row, 12, "")  # No advance for company
                ws.cell(row, 13, parsed_order.get("company_kms_travelled", 0))
                ws.cell(row, 14, parsed_order.get("company_toll", 0))
                ws.cell(row, 15, parsed_order.get("company_diesel", 0))
            
            ws.cell(row, 16, "Active")
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="EXPORT",
            resource_type="ORDER",
            new_data={"format": "excel", "count": len(orders)}
        )
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=kawale_cranes_orders.xlsx"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")

@api_router.get("/export/pdf")
async def export_orders_pdf(
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN])),
    order_type: Optional[str] = Query(None),
    customer_name: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    limit: int = Query(1000, ge=1, le=5000)
):
    """Export orders to PDF (Admin and Super Admin only)"""
    try:
        # Build query
        query = {}
        if order_type:
            query["order_type"] = order_type
        if customer_name:
            query["customer_name"] = {"$regex": customer_name, "$options": "i"}
        if phone:
            query["phone"] = {"$regex": phone, "$options": "i"}
        
        # Get orders
        orders = await db.crane_orders.find(query, {"_id": 0}).limit(limit).to_list(limit)
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#366092')
        )
        
        # Title
        story.append(Paragraph("Kawale Cranes - Orders Report", title_style))
        story.append(Spacer(1, 20))
        
        # Summary
        summary_data = [
            ['Total Orders', str(len(orders))],
            ['Report Generated', datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
            ['Generated By', current_user.get("full_name", "Admin")]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Orders table
        headers = ["ID", "Date", "Customer", "Phone", "Type", "Amount"]
        data = [headers]
        
        for order in orders:
            parsed_order = parse_from_mongo(order)
            row = [
                parsed_order.get("unique_id", "")[:8],
                parsed_order.get("date_time").strftime("%Y-%m-%d") if parsed_order.get("date_time") else "",
                parsed_order.get("customer_name", "")[:20],
                parsed_order.get("phone", ""),
                parsed_order.get("order_type", "").title(),
                f"₹{parsed_order.get('amount_received', 0)}" if parsed_order.get("order_type") == "cash" else "N/A"
            ]
            data.append(row)
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="EXPORT",
            resource_type="ORDER",
            new_data={"format": "pdf", "count": len(orders)}
        )
        
        return StreamingResponse(
            io.BytesIO(buffer.getvalue()),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=kawale_cranes_orders.pdf"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting PDF: {str(e)}")

# Financial calculation endpoint
@api_router.get("/orders/{order_id}/financials")
async def get_order_financials(
    order_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get calculated financial details for an order"""
    try:
        # Get the order
        order = await db.crane_orders.find_one({"id": order_id}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Calculate financials
        financials = await calculate_order_financials(order)
        
        return financials
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating financials: {str(e)}")

@api_router.get("/rates")
async def get_service_rates(
    current_user: dict = Depends(get_current_user)
):
    """Get all service rates (All authenticated users can view)"""
    try:
        rates = await db.service_rates.find({}, {"_id": 0}).to_list(1000)
        return [parse_from_mongo(rate) for rate in rates]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching service rates: {str(e)}")

@api_router.put("/rates/{rate_id}")
async def update_service_rate(
    rate_id: str,
    update_data: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Update a service rate (Admin and Super Admin only)"""
    try:
        # Validate the update data
        allowed_fields = ['base_rate', 'base_distance_km', 'rate_per_km_beyond']
        update_dict = {k: v for k, v in update_data.items() if k in allowed_fields}
        
        if not update_dict:
            raise HTTPException(status_code=400, detail="No valid fields to update")
        
        # Validate numeric values
        for field, value in update_dict.items():
            if not isinstance(value, (int, float)) or value < 0:
                raise HTTPException(status_code=400, detail=f"{field} must be a non-negative number")
        
        # Add audit fields
        update_dict['updated_at'] = datetime.now(timezone.utc)
        
        # Prepare for MongoDB
        prepared_update = prepare_for_mongo(update_dict)
        
        # Update the rate
        result = await db.service_rates.update_one(
            {"id": rate_id},
            {"$set": prepared_update}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Service rate not found")
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="UPDATE",
            resource_type="SERVICE_RATE",
            resource_id=rate_id,
            new_data=update_dict
        )
        
        # Return updated rate
        updated_rate = await db.service_rates.find_one({"id": rate_id}, {"_id": 0})
        return parse_from_mongo(updated_rate)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating service rate: {str(e)}")

@api_router.post("/rates")
async def create_service_rate(
    rate_data: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Create a new service rate (Admin and Super Admin only)"""
    try:
        # Validate required fields
        required_fields = ['name_of_firm', 'company_name', 'service_type', 'base_rate', 'rate_per_km_beyond']
        missing_fields = [field for field in required_fields if field not in rate_data or rate_data[field] is None]
        
        if missing_fields:
            field_list = ", ".join(missing_fields)
            raise HTTPException(status_code=400, detail=f"Missing required fields: {field_list}")
        
        # Check if rate combination already exists
        existing_rate = await db.service_rates.find_one({
            "name_of_firm": rate_data["name_of_firm"],
            "company_name": rate_data["company_name"],
            "service_type": rate_data["service_type"]
        })
        
        if existing_rate:
            raise HTTPException(
                status_code=400, 
                detail=f"Rate already exists for {rate_data['name_of_firm']} - {rate_data['company_name']} - {rate_data['service_type']}"
            )
        
        # Create new rate
        new_rate = ServiceRate(**rate_data)
        doc = prepare_for_mongo(new_rate.model_dump())
        
        result = await db.service_rates.insert_one(doc)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="CREATE",
            resource_type="SERVICE_RATE",
            resource_id=new_rate.id,
            new_data=rate_data
        )
        
        return new_rate
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating service rate: {str(e)}")

@api_router.delete("/rates/{rate_id}")
async def delete_service_rate(
    rate_id: str,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Delete a service rate (Admin and Super Admin only)"""
    try:
        # Get rate for audit log
        existing_rate = await db.service_rates.find_one({"id": rate_id}, {"_id": 0})
        if not existing_rate:
            raise HTTPException(status_code=404, detail="Service rate not found")
        
        result = await db.service_rates.delete_one({"id": rate_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Service rate not found")
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="DELETE",
            resource_type="SERVICE_RATE",
            resource_id=rate_id,
            old_data=existing_rate
        )
        
        return {"message": "Service rate deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting service rate: {str(e)}")



# Driver Salary Management endpoints
@api_router.get("/drivers/list")
async def get_all_drivers(
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Get list of all unique drivers from orders"""
    try:
        # Get unique driver names from both cash and company orders
        cash_drivers = await db.crane_orders.distinct("cash_driver_name")
        company_drivers = await db.crane_orders.distinct("company_driver_name")
        
        # Combine and remove None/empty values
        all_drivers = set()
        for driver in cash_drivers:
            if driver and str(driver).strip():
                all_drivers.add(str(driver).strip())
        for driver in company_drivers:
            if driver and str(driver).strip():
                all_drivers.add(str(driver).strip())
        
        # Get default salaries for these drivers
        default_salaries = await db.driver_default_salaries.find({}).to_list(length=None)
        salary_map = {ds["driver_name"]: ds["default_salary"] for ds in default_salaries}
        
        # Build response with default salary info
        drivers_list = []
        for driver in sorted(all_drivers):
            drivers_list.append({
                "name": driver,
                "default_salary": salary_map.get(driver, 15000.0)
            })
        
        return drivers_list
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching drivers: {str(e)}")

@api_router.post("/drivers/default-salary")
async def set_driver_default_salary(
    data: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Set default salary for a driver (applies to all months)"""
    try:
        driver_name = data.get("driver_name")
        default_salary = float(data.get("default_salary", 15000.0))
        
        if not driver_name:
            raise HTTPException(status_code=400, detail="Driver name is required")
        
        # Check if default salary exists
        existing = await db.driver_default_salaries.find_one({"driver_name": driver_name})
        
        if existing:
            # Update existing
            result = await db.driver_default_salaries.update_one(
                {"driver_name": driver_name},
                {
                    "$set": {
                        "default_salary": default_salary,
                        "updated_by": current_user["full_name"],
                        "updated_by_email": current_user["email"],
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            )
        else:
            # Create new
            default_salary_record = DriverDefaultSalary(
                driver_name=driver_name,
                default_salary=default_salary,
                updated_by=current_user["full_name"],
                updated_by_email=current_user["email"]
            )
            salary_dict = prepare_for_mongo(default_salary_record.model_dump())
            await db.driver_default_salaries.insert_one(salary_dict)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="UPDATE",
            resource_type="DRIVER_DEFAULT_SALARY",
            resource_id=driver_name,
            new_data={"driver_name": driver_name, "default_salary": default_salary}
        )
        
        return {"message": f"Default salary set to ₹{default_salary} for {driver_name}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error setting default salary: {str(e)}")

@api_router.post("/drivers/bulk-default-salary")
async def set_bulk_default_salary(
    data: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Set default salary for multiple drivers at once"""
    try:
        drivers = data.get("drivers", [])  # [{name, default_salary}]
        
        if not drivers:
            raise HTTPException(status_code=400, detail="Drivers list is required")
        
        updated_count = 0
        for driver_info in drivers:
            driver_name = driver_info.get("name")
            default_salary = float(driver_info.get("default_salary", 15000.0))
            
            if not driver_name:
                continue
            
            existing = await db.driver_default_salaries.find_one({"driver_name": driver_name})
            
            if existing:
                await db.driver_default_salaries.update_one(
                    {"driver_name": driver_name},
                    {
                        "$set": {
                            "default_salary": default_salary,
                            "updated_by": current_user["full_name"],
                            "updated_by_email": current_user["email"],
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
            else:
                default_salary_record = DriverDefaultSalary(
                    driver_name=driver_name,
                    default_salary=default_salary,
                    updated_by=current_user["full_name"],
                    updated_by_email=current_user["email"]
                )
                salary_dict = prepare_for_mongo(default_salary_record.model_dump())
                await db.driver_default_salaries.insert_one(salary_dict)
            
            updated_count += 1
        
        return {"message": f"Default salaries updated for {updated_count} drivers"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error setting bulk default salaries: {str(e)}")

@api_router.get("/driver-salaries")
async def get_driver_salaries(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2020, le=2030),
    driver_name: Optional[str] = None,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Get driver salaries with optional filters"""
    try:
        query = {}
        if month:
            query["month"] = month
        if year:
            query["year"] = year
        if driver_name:
            query["driver_name"] = driver_name
        
        salaries = await db.driver_salaries.find(query).sort([("year", -1), ("month", -1)]).to_list(length=None)
        
        # Convert datetime fields
        for salary in salaries:
            if isinstance(salary.get('added_at'), datetime):
                salary['added_at'] = salary['added_at'].isoformat()
            if isinstance(salary.get('updated_at'), datetime):
                salary['updated_at'] = salary['updated_at'].isoformat()
        
        return salaries
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching driver salaries: {str(e)}")

@api_router.post("/driver-salaries")
async def create_driver_salary(
    salary_data: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Create new driver salary record"""
    try:
        # Check if salary already exists for this driver/month/year
        existing = await db.driver_salaries.find_one({
            "driver_name": salary_data.get("driver_name"),
            "month": salary_data.get("month"),
            "year": salary_data.get("year")
        })
        
        if existing:
            raise HTTPException(status_code=400, detail="Salary record already exists for this driver/month/year")
        
        # Calculate total incentives for the month
        start_date = datetime(salary_data["year"], salary_data["month"], 1, tzinfo=timezone.utc)
        if salary_data["month"] == 12:
            end_date = datetime(salary_data["year"] + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(salary_data["year"], salary_data["month"] + 1, 1, tzinfo=timezone.utc)
        
        # Get incentives from orders
        incentives_pipeline = [
            {
                "$match": {
                    "$or": [
                        {"cash_driver_name": salary_data.get("driver_name")},
                        {"company_driver_name": salary_data.get("driver_name")}
                    ],
                    "date_time": {
                        "$gte": start_date.isoformat(),
                        "$lt": end_date.isoformat()
                    },
                    "incentive_amount": {"$ne": None, "$gt": 0}
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total_incentives": {"$sum": "$incentive_amount"}
                }
            }
        ]
        
        incentives_result = await db.crane_orders.aggregate(incentives_pipeline).to_list(length=None)
        total_incentives = incentives_result[0]["total_incentives"] if incentives_result else 0.0
        
        salary = DriverSalary(
            driver_name=salary_data["driver_name"],
            month=salary_data["month"],
            year=salary_data["year"],
            base_salary=float(salary_data["base_salary"]),
            total_incentives=float(total_incentives),
            deductions=float(salary_data.get("deductions", 0.0)),
            notes=salary_data.get("notes"),
            added_by=current_user["full_name"],
            added_by_email=current_user["email"]
        )
        
        salary_dict = prepare_for_mongo(salary.model_dump())
        await db.driver_salaries.insert_one(salary_dict)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="CREATE",
            resource_type="DRIVER_SALARY",
            resource_id=salary.id,
            new_data=salary.model_dump(mode='json')
        )
        
        return salary
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating driver salary: {str(e)}")

@api_router.put("/driver-salaries/{salary_id}")
async def update_driver_salary(
    salary_id: str,
    salary_data: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Update driver salary record"""
    try:
        existing = await db.driver_salaries.find_one({"id": salary_id}, {"_id": 0})
        if not existing:
            raise HTTPException(status_code=404, detail="Salary record not found")
        
        # Recalculate incentives if month/year/driver changed
        month = salary_data.get("month", existing["month"])
        year = salary_data.get("year", existing["year"])
        driver_name = salary_data.get("driver_name", existing["driver_name"])
        
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        
        incentives_pipeline = [
            {
                "$match": {
                    "$or": [
                        {"cash_driver_name": driver_name},
                        {"company_driver_name": driver_name}
                    ],
                    "date_time": {
                        "$gte": start_date.isoformat(),
                        "$lt": end_date.isoformat()
                    },
                    "incentive_amount": {"$ne": None, "$gt": 0}
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total_incentives": {"$sum": "$incentive_amount"}
                }
            }
        ]
        
        incentives_result = await db.crane_orders.aggregate(incentives_pipeline).to_list(length=None)
        total_incentives = incentives_result[0]["total_incentives"] if incentives_result else 0.0
        
        update_data = {
            "base_salary": float(salary_data.get("base_salary", existing["base_salary"])),
            "total_incentives": float(total_incentives),
            "deductions": float(salary_data.get("deductions", existing.get("deductions", 0.0))),
            "notes": salary_data.get("notes", existing.get("notes")),
            "updated_by": current_user["full_name"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = await db.driver_salaries.update_one(
            {"id": salary_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Salary record not found")
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="UPDATE",
            resource_type="DRIVER_SALARY",
            resource_id=salary_id,
            old_data=existing,
            new_data=update_data
        )
        
        return {"message": "Driver salary updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating driver salary: {str(e)}")

@api_router.delete("/driver-salaries/{salary_id}")
async def delete_driver_salary(
    salary_id: str,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Delete driver salary record"""
    try:
        existing = await db.driver_salaries.find_one({"id": salary_id}, {"_id": 0})
        if not existing:
            raise HTTPException(status_code=404, detail="Salary record not found")
        
        result = await db.driver_salaries.delete_one({"id": salary_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Salary record not found")
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="DELETE",
            resource_type="DRIVER_SALARY",
            resource_id=salary_id,
            old_data=existing
        )
        
        return {"message": "Driver salary deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting driver salary: {str(e)}")

@api_router.get("/driver-salaries/calculate-incentives")
async def calculate_driver_incentives(
    driver_name: str = Query(...),
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2030),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Calculate total incentives for a driver in a specific month"""
    try:
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        
        # Get all orders with incentives for this driver
        orders = await db.crane_orders.find({
            "$or": [
                {"cash_driver_name": driver_name},
                {"company_driver_name": driver_name}
            ],
            "date_time": {
                "$gte": start_date.isoformat(),
                "$lt": end_date.isoformat()
            },
            "incentive_amount": {"$ne": None, "$gt": 0}
        }, {"_id": 0, "unique_id": 1, "customer_name": 1, "date_time": 1, "incentive_amount": 1, "incentive_reason": 1, "order_type": 1}).to_list(length=None)
        
        total_incentives = sum(order.get("incentive_amount", 0) for order in orders)
        
        return {
            "driver_name": driver_name,
            "month": month,
            "year": year,
            "total_incentives": total_incentives,
            "incentive_count": len(orders),
            "orders": orders
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating incentives: {str(e)}")

# Reports endpoints
@api_router.get("/reports/expense-by-driver")
async def get_expense_report_by_driver(
    month: int = Query(..., ge=1, le=12, description="Month (1-12)"),
    year: int = Query(..., ge=2020, le=2030, description="Year (2020-2030)"),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Get expense report by driver for a specific month (Admin and Super Admin only)"""
    try:
        # Create date range for the month
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        
        # Query orders for the specified month
        query = {
            "date_time": {
                "$gte": start_date.isoformat(),
                "$lt": end_date.isoformat()
            }
        }
        
        orders = await db.crane_orders.find(query, {"_id": 0}).to_list(10000)
        parsed_orders = [parse_from_mongo(order) for order in orders]
        
        # Aggregate expenses by driver
        driver_expenses = {}
        
        for order in parsed_orders:
            # Determine driver name based on order type
            if order.get("order_type") == "cash":
                driver = order.get("cash_driver_name") or "Unknown Driver"
                diesel_cost = order.get("cash_diesel", 0) or 0
                toll_cost = order.get("cash_toll", 0) or 0
            else:  # company order
                driver = order.get("company_driver_name") or "Unknown Driver"
                diesel_cost = order.get("company_diesel", 0) or 0
                toll_cost = order.get("company_toll", 0) or 0
            
            if driver not in driver_expenses:
                driver_expenses[driver] = {
                    "driver_name": driver,
                    "cash_orders": 0,
                    "company_orders": 0,
                    "total_orders": 0,
                    "total_diesel_expense": 0,
                    "total_toll_expense": 0,
                    "total_expenses": 0
                }
            
            # Update counts
            if order.get("order_type") == "cash":
                driver_expenses[driver]["cash_orders"] += 1
            else:
                driver_expenses[driver]["company_orders"] += 1
            
            driver_expenses[driver]["total_orders"] += 1
            driver_expenses[driver]["total_diesel_expense"] += diesel_cost
            driver_expenses[driver]["total_toll_expense"] += toll_cost
            driver_expenses[driver]["total_expenses"] = (
                driver_expenses[driver]["total_diesel_expense"] + 
                driver_expenses[driver]["total_toll_expense"]
            )
        
        # Convert to list and sort by total expenses
        report_data = list(driver_expenses.values())
        report_data.sort(key=lambda x: x["total_expenses"], reverse=True)
        
        return {
            "month": month,
            "year": year,
            "report_type": "expense_by_driver",
            "data": report_data,
            "summary": {
                "total_drivers": len(report_data),
                "total_orders": sum(d["total_orders"] for d in report_data),
                "total_expenses": sum(d["total_expenses"] for d in report_data)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating expense report: {str(e)}")

@api_router.get("/reports/revenue-by-towing-vehicle")
async def get_revenue_report_by_towing_vehicle(
    month: int = Query(..., ge=1, le=12, description="Month (1-12)"),
    year: int = Query(..., ge=2020, le=2030, description="Year (2020-2030)"),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Get revenue report by towing vehicle for a specific month (Admin and Super Admin only)"""
    try:
        # Create date range for the month
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        
        # Query orders for the specified month
        query = {
            "date_time": {
                "$gte": start_date.isoformat(),
                "$lt": end_date.isoformat()
            }
        }
        
        orders = await db.crane_orders.find(query, {"_id": 0}).to_list(10000)
        parsed_orders = [parse_from_mongo(order) for order in orders]
        
        # Aggregate revenue by towing vehicle
        vehicle_revenue = {}
        
        for order in parsed_orders:
            # Determine towing vehicle based on order type
            if order.get("order_type") == "cash":
                towing_vehicle = order.get("cash_towing_vehicle") or "Unknown Vehicle"
                # For cash orders, revenue is amount_received
                base_revenue = order.get("amount_received", 0) or 0
            else:  # company order
                towing_vehicle = order.get("company_towing_vehicle") or "Unknown Vehicle"
                # For company orders, calculate revenue using rates
                financials = await calculate_order_financials(order)
                base_revenue = financials.base_revenue
            
            # Add incentive to revenue
            incentive_amount = order.get("incentive_amount", 0) or 0
            total_revenue = base_revenue + incentive_amount
            
            if towing_vehicle not in vehicle_revenue:
                vehicle_revenue[towing_vehicle] = {
                    "towing_vehicle": towing_vehicle,
                    "cash_orders": 0,
                    "company_orders": 0,
                    "total_orders": 0,
                    "total_base_revenue": 0,
                    "total_incentive_amount": 0,
                    "total_revenue": 0
                }
            
            # Update counts and revenue
            if order.get("order_type") == "cash":
                vehicle_revenue[towing_vehicle]["cash_orders"] += 1
            else:
                vehicle_revenue[towing_vehicle]["company_orders"] += 1
            
            vehicle_revenue[towing_vehicle]["total_orders"] += 1
            vehicle_revenue[towing_vehicle]["total_base_revenue"] += base_revenue
            vehicle_revenue[towing_vehicle]["total_incentive_amount"] += incentive_amount
            vehicle_revenue[towing_vehicle]["total_revenue"] += total_revenue
        
        # Convert to list and sort by total revenue
        report_data = list(vehicle_revenue.values())
        report_data.sort(key=lambda x: x["total_revenue"], reverse=True)
        
        return {
            "month": month,
            "year": year,
            "report_type": "revenue_by_towing_vehicle",
            "data": report_data,
            "summary": {
                "total_vehicles": len(report_data),
                "total_orders": sum(d["total_orders"] for d in report_data),
                "total_revenue": sum(d["total_revenue"] for d in report_data)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating towing vehicle revenue report: {str(e)}")

@api_router.get("/reports/revenue-by-vehicle-type")
async def get_revenue_report_by_vehicle_type(
    month: int = Query(..., ge=1, le=12, description="Month (1-12)"),
    year: int = Query(..., ge=2020, le=2030, description="Year (2020-2030)"),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Get revenue report by vehicle type for a specific month (Admin and Super Admin only)"""
    try:
        # Create date range for the month
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        
        # Query orders for the specified month
        query = {
            "date_time": {
                "$gte": start_date.isoformat(),
                "$lt": end_date.isoformat()
            }
        }
        
        orders = await db.crane_orders.find(query, {"_id": 0}).to_list(10000)
        parsed_orders = [parse_from_mongo(order) for order in orders]
        
        # Aggregate revenue by vehicle type (service type)
        vehicle_revenue = {}
        
        for order in parsed_orders:
            # Determine service type based on order type
            if order.get("order_type") == "cash":
                service_type = order.get("cash_service_type") or "Unknown Service"
                # For cash orders, revenue is amount_received
                base_revenue = order.get("amount_received", 0) or 0
            else:  # company order
                service_type = order.get("company_service_type") or "Unknown Service"
                # For company orders, calculate revenue using rates
                financials = await calculate_order_financials(order)
                base_revenue = financials.base_revenue
            
            # Add incentive to revenue
            incentive_amount = order.get("incentive_amount", 0) or 0
            total_revenue = base_revenue + incentive_amount
            
            if service_type not in vehicle_revenue:
                vehicle_revenue[service_type] = {
                    "service_type": service_type,
                    "cash_orders": 0,
                    "company_orders": 0,
                    "total_orders": 0,
                    "total_base_revenue": 0,
                    "total_incentive_amount": 0,
                    "total_revenue": 0
                }
            
            # Update counts and revenue
            if order.get("order_type") == "cash":
                vehicle_revenue[service_type]["cash_orders"] += 1
            else:
                vehicle_revenue[service_type]["company_orders"] += 1
            
            vehicle_revenue[service_type]["total_orders"] += 1
            vehicle_revenue[service_type]["total_base_revenue"] += base_revenue
            vehicle_revenue[service_type]["total_incentive_amount"] += incentive_amount
            vehicle_revenue[service_type]["total_revenue"] += total_revenue
        
        # Convert to list and sort by total revenue
        report_data = list(vehicle_revenue.values())
        report_data.sort(key=lambda x: x["total_revenue"], reverse=True)
        
        return {
            "month": month,
            "year": year,
            "report_type": "revenue_by_vehicle_type",
            "data": report_data,
            "summary": {
                "total_service_types": len(report_data),
                "total_orders": sum(d["total_orders"] for d in report_data),
                "total_revenue": sum(d["total_revenue"] for d in report_data)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating revenue report: {str(e)}")

@api_router.get("/reports/expense-by-driver/export")
async def export_expense_report_by_driver(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2030),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Export expense report by driver as Excel file"""
    try:
        # Get the report data
        report_response = await get_expense_report_by_driver(month, year, current_user)
        report_data = report_response["data"]
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"Driver Expenses {year}-{month:02d}"
        
        # Headers
        headers = [
            "Driver Name", "Cash Orders", "Company Orders", "Total Orders",
            "Diesel Expenses (₹)", "Toll Expenses (₹)", "Total Expenses (₹)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_idx, driver_data in enumerate(report_data, 2):
            worksheet.cell(row=row_idx, column=1, value=driver_data["driver_name"])
            worksheet.cell(row=row_idx, column=2, value=driver_data["cash_orders"])
            worksheet.cell(row=row_idx, column=3, value=driver_data["company_orders"])
            worksheet.cell(row=row_idx, column=4, value=driver_data["total_orders"])
            worksheet.cell(row=row_idx, column=5, value=driver_data["total_diesel_expense"])
            worksheet.cell(row=row_idx, column=6, value=driver_data["total_toll_expense"])
            worksheet.cell(row=row_idx, column=7, value=driver_data["total_expenses"])
        
        # Summary row
        summary_row = len(report_data) + 3
        worksheet.cell(row=summary_row, column=1, value="TOTAL")
        worksheet.cell(row=summary_row, column=2, value=sum(d["cash_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=3, value=sum(d["company_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=4, value=sum(d["total_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=5, value=sum(d["total_diesel_expense"] for d in report_data))
        worksheet.cell(row=summary_row, column=6, value=sum(d["total_toll_expense"] for d in report_data))
        worksheet.cell(row=summary_row, column=7, value=sum(d["total_expenses"] for d in report_data))
        
        # Format summary row
        for col in range(1, 8):
            cell = worksheet.cell(row=summary_row, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"kawale_expense_by_driver_{year}_{month:02d}.xlsx"
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting expense report: {str(e)}")

@api_router.get("/reports/revenue-by-towing-vehicle/export")
async def export_revenue_report_by_towing_vehicle(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2030),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Export revenue report by towing vehicle as Excel file"""
    try:
        # Get the report data
        report_response = await get_revenue_report_by_towing_vehicle(month, year, current_user)
        report_data = report_response["data"]
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"Towing Vehicle Revenue {year}-{month:02d}"
        
        # Headers
        headers = [
            "Towing Vehicle", "Cash Orders", "Company Orders", "Total Orders",
            "Base Revenue (₹)", "Incentive Amount (₹)", "Total Revenue (₹)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_idx, vehicle_data in enumerate(report_data, 2):
            worksheet.cell(row=row_idx, column=1, value=vehicle_data["towing_vehicle"])
            worksheet.cell(row=row_idx, column=2, value=vehicle_data["cash_orders"])
            worksheet.cell(row=row_idx, column=3, value=vehicle_data["company_orders"])
            worksheet.cell(row=row_idx, column=4, value=vehicle_data["total_orders"])
            worksheet.cell(row=row_idx, column=5, value=vehicle_data["total_base_revenue"])
            worksheet.cell(row=row_idx, column=6, value=vehicle_data["total_incentive_amount"])
            worksheet.cell(row=row_idx, column=7, value=vehicle_data["total_revenue"])
        
        # Summary row
        summary_row = len(report_data) + 3
        worksheet.cell(row=summary_row, column=1, value="TOTAL")
        worksheet.cell(row=summary_row, column=2, value=sum(d["cash_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=3, value=sum(d["company_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=4, value=sum(d["total_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=5, value=sum(d["total_base_revenue"] for d in report_data))
        worksheet.cell(row=summary_row, column=6, value=sum(d["total_incentive_amount"] for d in report_data))
        worksheet.cell(row=summary_row, column=7, value=sum(d["total_revenue"] for d in report_data))
        
        # Format summary row
        for col in range(1, 8):
            cell = worksheet.cell(row=summary_row, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"kawale_revenue_by_towing_vehicle_{year}_{month:02d}.xlsx"
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting towing vehicle revenue report: {str(e)}")

@api_router.get("/reports/revenue-by-vehicle-type/export")
async def export_revenue_report_by_vehicle_type(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2030),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Export revenue report by vehicle type as Excel file"""
    try:
        # Get the report data
        report_response = await get_revenue_report_by_vehicle_type(month, year, current_user)
        report_data = report_response["data"]
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"Vehicle Revenue {year}-{month:02d}"
        
        # Headers
        headers = [
            "Service Type", "Cash Orders", "Company Orders", "Total Orders",
            "Base Revenue (₹)", "Incentive Amount (₹)", "Total Revenue (₹)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_idx, vehicle_data in enumerate(report_data, 2):
            worksheet.cell(row=row_idx, column=1, value=vehicle_data["service_type"])
            worksheet.cell(row=row_idx, column=2, value=vehicle_data["cash_orders"])
            worksheet.cell(row=row_idx, column=3, value=vehicle_data["company_orders"])
            worksheet.cell(row=row_idx, column=4, value=vehicle_data["total_orders"])
            worksheet.cell(row=row_idx, column=5, value=vehicle_data["total_base_revenue"])
            worksheet.cell(row=row_idx, column=6, value=vehicle_data["total_incentive_amount"])
            worksheet.cell(row=row_idx, column=7, value=vehicle_data["total_revenue"])
        
        # Summary row
        summary_row = len(report_data) + 3
        worksheet.cell(row=summary_row, column=1, value="TOTAL")
        worksheet.cell(row=summary_row, column=2, value=sum(d["cash_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=3, value=sum(d["company_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=4, value=sum(d["total_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=5, value=sum(d["total_base_revenue"] for d in report_data))
        worksheet.cell(row=summary_row, column=6, value=sum(d["total_incentive_amount"] for d in report_data))
        worksheet.cell(row=summary_row, column=7, value=sum(d["total_revenue"] for d in report_data))
        
        # Format summary row
        for col in range(1, 8):
            cell = worksheet.cell(row=summary_row, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"kawale_revenue_by_vehicle_type_{year}_{month:02d}.xlsx"
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting revenue report: {str(e)}")

@api_router.post("/reports/custom")
async def generate_custom_report(
    report_config: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Generate custom report based on user configuration"""
    try:
        # Extract configuration
        start_date_str = report_config.get("start_date")
        end_date_str = report_config.get("end_date")
        group_by = report_config.get("group_by", "order_type")  # order_type, driver, service_type, towing_vehicle, firm, company
        report_type = report_config.get("report_type", "summary")  # summary, detailed
        order_types = report_config.get("order_types", ["cash", "company"])  # filter by order types
        
        # Parse dates
        start_date = datetime.fromisoformat(start_date_str.replace('Z', '+00:00')) if start_date_str else datetime(2020, 1, 1, tzinfo=timezone.utc)
        end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')) if end_date_str else datetime.now(timezone.utc)
        
        # Query orders for the date range
        query = {
            "date_time": {
                "$gte": start_date.isoformat(),
                "$lte": end_date.isoformat()
            }
        }
        
        if order_types and len(order_types) < 2:
            query["order_type"] = {"$in": order_types}
        
        orders = await db.crane_orders.find(query, {"_id": 0}).to_list(10000)
        parsed_orders = [parse_from_mongo(order) for order in orders]
        
        # Group data based on configuration
        grouped_data = {}
        
        for order in parsed_orders:
            # Determine grouping key
            if group_by == "driver":
                key = order.get("cash_driver_name" if order.get("order_type") == "cash" else "company_driver_name") or "Unknown Driver"
            elif group_by == "service_type":
                key = order.get("cash_service_type" if order.get("order_type") == "cash" else "company_service_type") or "Unknown Service"
            elif group_by == "towing_vehicle":
                key = order.get("cash_towing_vehicle" if order.get("order_type") == "cash" else "company_towing_vehicle") or "Unknown Vehicle"
            elif group_by == "firm":
                key = order.get("name_of_firm") or "Unknown Firm"
            elif group_by == "company":
                key = order.get("company_name") or "Unknown Company"
            else:  # order_type
                key = order.get("order_type", "unknown").title()
            
            if key not in grouped_data:
                grouped_data[key] = {
                    "group_key": key,
                    "cash_orders": 0,
                    "company_orders": 0,
                    "total_orders": 0,
                    "total_revenue": 0,
                    "total_expenses": 0,
                    "total_incentives": 0,
                    "orders": [] if report_type == "detailed" else None
                }
            
            # Calculate revenue and expenses
            if order.get("order_type") == "cash":
                revenue = order.get("amount_received", 0) or 0
                expenses = (order.get("cash_diesel", 0) or 0) + (order.get("cash_toll", 0) or 0)
                grouped_data[key]["cash_orders"] += 1
            else:  # company order
                financials = await calculate_order_financials(order)
                revenue = financials.base_revenue
                expenses = (order.get("company_diesel", 0) or 0) + (order.get("company_toll", 0) or 0)
                grouped_data[key]["company_orders"] += 1
            
            incentive = order.get("incentive_amount", 0) or 0
            
            # Update aggregates
            grouped_data[key]["total_orders"] += 1
            grouped_data[key]["total_revenue"] += revenue + incentive
            grouped_data[key]["total_expenses"] += expenses
            grouped_data[key]["total_incentives"] += incentive
            
            # Add to detailed orders if needed
            if report_type == "detailed":
                grouped_data[key]["orders"].append({
                    "id": order.get("id"),
                    "customer_name": order.get("customer_name"),
                    "phone": order.get("phone"),
                    "order_type": order.get("order_type"),
                    "date_time": order.get("date_time").isoformat() if order.get("date_time") else None,
                    "revenue": revenue + incentive,
                    "expenses": expenses
                })
        
        # Convert to list and sort by total revenue
        report_data = list(grouped_data.values())
        report_data.sort(key=lambda x: x["total_revenue"], reverse=True)
        
        return {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "group_by": group_by,
            "report_type": report_type,
            "order_types": order_types,
            "data": report_data,
            "summary": {
                "total_groups": len(report_data),
                "total_orders": sum(d["total_orders"] for d in report_data),
                "total_revenue": sum(d["total_revenue"] for d in report_data),
                "total_expenses": sum(d["total_expenses"] for d in report_data)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating custom report: {str(e)}")

@api_router.post("/reports/custom/export")
async def export_custom_report(
    report_config: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Export custom report as Excel file"""
    try:
        # Get the report data
        report_response = await generate_custom_report(report_config, current_user)
        report_data = report_response["data"]
        group_by = report_response["group_by"]
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"Custom Report by {group_by.replace('_', ' ').title()}"
        
        # Headers
        headers = [
            group_by.replace('_', ' ').title(),
            "Cash Orders", "Company Orders", "Total Orders",
            "Total Revenue (₹)", "Total Expenses (₹)", "Total Incentives (₹)", "Net Profit (₹)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_idx, group_data in enumerate(report_data, 2):
            net_profit = group_data["total_revenue"] - group_data["total_expenses"]
            worksheet.cell(row=row_idx, column=1, value=group_data["group_key"])
            worksheet.cell(row=row_idx, column=2, value=group_data["cash_orders"])
            worksheet.cell(row=row_idx, column=3, value=group_data["company_orders"])
            worksheet.cell(row=row_idx, column=4, value=group_data["total_orders"])
            worksheet.cell(row=row_idx, column=5, value=group_data["total_revenue"])
            worksheet.cell(row=row_idx, column=6, value=group_data["total_expenses"])
            worksheet.cell(row=row_idx, column=7, value=group_data["total_incentives"])
            worksheet.cell(row=row_idx, column=8, value=net_profit)
        
        # Summary row
        summary_row = len(report_data) + 3
        total_revenue = sum(d["total_revenue"] for d in report_data)
        total_expenses = sum(d["total_expenses"] for d in report_data)
        worksheet.cell(row=summary_row, column=1, value="TOTAL")
        worksheet.cell(row=summary_row, column=2, value=sum(d["cash_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=3, value=sum(d["company_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=4, value=sum(d["total_orders"] for d in report_data))
        worksheet.cell(row=summary_row, column=5, value=total_revenue)
        worksheet.cell(row=summary_row, column=6, value=total_expenses)
        worksheet.cell(row=summary_row, column=7, value=sum(d["total_incentives"] for d in report_data))
        worksheet.cell(row=summary_row, column=8, value=total_revenue - total_expenses)
        
        # Format summary row
        for col in range(1, 9):
            cell = worksheet.cell(row=summary_row, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"kawale_custom_report_{group_by}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting custom report: {str(e)}")



@api_router.get("/import/history")
async def get_import_history(
    limit: int = Query(10, ge=1, le=100),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Get import history (Admin and Super Admin only)"""
    try:
        imports = await db.import_history.find({}, {"_id": 0}).sort("imported_at", -1).limit(limit).to_list(length=None)
        
        # Convert datetime fields to ISO format
        for import_record in imports:
            if isinstance(import_record.get('imported_at'), datetime):
                import_record['imported_at'] = import_record['imported_at'].isoformat()
        
        return imports
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching import history: {str(e)}")

@api_router.post("/import/save-history")
async def save_import_history(
    import_data: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Save import history metadata"""
    try:
        import_history = ImportHistory(
            filename=import_data.get("filename", "unknown"),
            imported_by=current_user["full_name"],
            imported_by_email=current_user["email"],
            total_records=import_data.get("total_records", 0),
            success_count=import_data.get("success_count", 0),
            error_count=import_data.get("error_count", 0),
            cash_orders=import_data.get("cash_orders", 0),
            company_orders=import_data.get("company_orders", 0),
            sample_data=import_data.get("sample_data", [])
        )
        
        import_dict = prepare_for_mongo(import_history.model_dump())
        result = await db.import_history.insert_one(import_dict)
        
        return {"message": "Import history saved successfully", "id": import_history.id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving import history: {str(e)}")

        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting custom report: {str(e)}")



@api_router.get("/reports/daily-summary")
async def get_daily_summary(
    start_date: str = Query(...),
    end_date: str = Query(...),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Get daily expense and revenue summary"""
    try:
        # Parse dates
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # Query orders in date range
        orders = await db.crane_orders.find({
            "date_time": {
                "$gte": start.isoformat(),
                "$lte": end.isoformat()
            }
        }, {"_id": 0}).to_list(10000)
        
        # Group by date
        daily_data = {}
        
        for order in orders:
            date_str = order.get('date_time', '')[:10]  # Get YYYY-MM-DD
            
            if date_str not in daily_data:
                daily_data[date_str] = {
                    'date': date_str,
                    'total_orders': 0,
                    'cash_orders': 0,
                    'company_orders': 0,
                    'total_expense': 0.0,
                    'total_revenue': 0.0,
                    'cash_revenue': 0.0,
                    'company_revenue': 0.0
                }
            
            daily_data[date_str]['total_orders'] += 1
            
            if order.get('order_type') == 'cash':
                daily_data[date_str]['cash_orders'] += 1
                # Cash expenses (diesel, toll, etc.)
                cash_diesel = order.get('cash_diesel', 0) or 0
                cash_toll = order.get('cash_toll', 0) or 0
                expense = float(cash_diesel) + float(cash_toll)
                daily_data[date_str]['total_expense'] += expense
                
                # Cash revenue
                amount = order.get('amount_received', 0) or 0
                daily_data[date_str]['cash_revenue'] += float(amount)
                daily_data[date_str]['total_revenue'] += float(amount)
                
            elif order.get('order_type') == 'company':
                daily_data[date_str]['company_orders'] += 1
                # Company expenses
                company_diesel = order.get('company_diesel', 0) or 0
                company_toll = order.get('company_toll', 0) or 0
                expense = float(company_diesel) + float(company_toll)
                daily_data[date_str]['total_expense'] += expense
                
                # Company revenue (calculated from rates if available)
                # For now, use a placeholder calculation
                # This can be enhanced with actual rate calculation
                company_revenue = 0.0
                # You can call calculate_financials here if needed
                daily_data[date_str]['company_revenue'] += company_revenue
                daily_data[date_str]['total_revenue'] += company_revenue
        
        # Convert to sorted list
        summary = sorted(daily_data.values(), key=lambda x: x['date'])
        
        return {
            "summary": summary,
            "totals": {
                "total_orders": sum(d['total_orders'] for d in summary),
                "total_expense": sum(d['total_expense'] for d in summary),
                "total_revenue": sum(d['total_revenue'] for d in summary),
                "net_profit": sum(d['total_revenue'] - d['total_expense'] for d in summary)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating daily summary: {str(e)}")

@api_router.post("/reports/custom-columns")
async def get_custom_column_report(
    report_config: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Generate fully custom report with selected columns"""
    try:
        start_date_str = report_config.get("start_date")
        end_date_str = report_config.get("end_date")
        selected_columns = report_config.get("columns", [])
        order_type_filter = report_config.get("order_type", "all")
        
        if not selected_columns:
            raise HTTPException(status_code=400, detail="At least one column must be selected")
        
        # Parse dates
        start = datetime.fromisoformat(start_date_str.replace('Z', '+00:00')) if start_date_str else datetime(2020, 1, 1, tzinfo=timezone.utc)
        end = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')) if end_date_str else datetime.now(timezone.utc)
        
        # Build query
        query = {
            "date_time": {
                "$gte": start.isoformat(),
                "$lte": end.isoformat()
            }
        }
        
        if order_type_filter != "all":
            query["order_type"] = order_type_filter
        
        # Build projection for selected columns
        projection = {"_id": 0, "id": 1}
        for col in selected_columns:
            projection[col] = 1
        
        # Fetch orders with only selected columns
        orders = await db.crane_orders.find(query, projection).to_list(10000)
        
        # Parse and format data
        formatted_orders = []
        for order in orders:
            formatted_order = {}
            for col in selected_columns:
                value = order.get(col)
                # Format based on column type
                if col in ['amount_received', 'advance_amount', 'cash_toll', 'company_toll', 
                          'cash_diesel', 'company_diesel', 'incentive_amount', 'care_off_amount',
                          'base_rate', 'total_expense', 'total_revenue']:
                    formatted_order[col] = float(value) if value else 0.0
                elif col in ['date_time', 'added_time', 'reach_time', 'drop_time']:
                    formatted_order[col] = value if value else None
                else:
                    formatted_order[col] = value if value else ''
            
            formatted_orders.append(formatted_order)
        
        return {
            "data": formatted_orders,
            "columns": selected_columns,
            "total_records": len(formatted_orders),
            "date_range": {
                "start": start.isoformat(),
                "end": end.isoformat()
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating custom report: {str(e)}")


@api_router.post("/reports/custom-columns/export/excel")
async def export_custom_columns_excel(
    report_config: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Export custom column report to Excel"""
    try:
        start_date_str = report_config.get("start_date")
        end_date_str = report_config.get("end_date")
        selected_columns = report_config.get("columns", [])
        order_type_filter = report_config.get("order_type", "all")
        
        if not selected_columns:
            raise HTTPException(status_code=400, detail="At least one column must be selected")
        
        # Parse dates
        start = datetime.fromisoformat(start_date_str.replace('Z', '+00:00')) if start_date_str else datetime(2020, 1, 1, tzinfo=timezone.utc)
        end = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')) if end_date_str else datetime.now(timezone.utc)
        
        # Build query
        query = {
            "date_time": {
                "$gte": start.isoformat(),
                "$lte": end.isoformat()
            }
        }
        
        if order_type_filter != "all":
            query["order_type"] = order_type_filter
        
        # Build projection for selected columns
        projection = {"_id": 0, "id": 1}
        for col in selected_columns:
            projection[col] = 1
        
        # Fetch orders with only selected columns
        orders = await db.crane_orders.find(query, projection).to_list(10000)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Custom Report"
        
        # Get column labels
        available_columns = [
            {"key": "unique_id", "label": "Order ID"},
            {"key": "date_time", "label": "Date & Time"},
            {"key": "customer_name", "label": "Customer Name"},
            {"key": "phone", "label": "Phone"},
            {"key": "order_type", "label": "Order Type"},
            {"key": "created_by", "label": "Created By"},
            {"key": "cash_trip_from", "label": "Trip From (Cash)"},
            {"key": "cash_trip_to", "label": "Trip To (Cash)"},
            {"key": "cash_driver_name", "label": "Driver (Cash)"},
            {"key": "cash_towing_vehicle", "label": "Towing Vehicle (Cash)"},
            {"key": "cash_service_type", "label": "Service Type (Cash)"},
            {"key": "cash_vehicle_name", "label": "Vehicle Name (Cash)"},
            {"key": "cash_vehicle_number", "label": "Vehicle Number (Cash)"},
            {"key": "amount_received", "label": "Amount Received"},
            {"key": "advance_amount", "label": "Advance Amount"},
            {"key": "cash_kms_travelled", "label": "KMs Travelled (Cash)"},
            {"key": "cash_toll", "label": "Toll (Cash)"},
            {"key": "cash_diesel", "label": "Diesel (Cash)"},
            {"key": "cash_diesel_refill_location", "label": "Diesel Refill Location"},
            {"key": "care_off", "label": "Care Off"},
            {"key": "care_off_amount", "label": "Care Off Amount"},
            {"key": "name_of_firm", "label": "Name of Firm"},
            {"key": "company_name", "label": "Company Name"},
            {"key": "case_id_file_number", "label": "Case ID/File Number"},
            {"key": "company_trip_from", "label": "Trip From (Company)"},
            {"key": "company_trip_to", "label": "Trip To (Company)"},
            {"key": "company_driver_name", "label": "Driver (Company)"},
            {"key": "company_towing_vehicle", "label": "Towing Vehicle (Company)"},
            {"key": "company_service_type", "label": "Service Type (Company)"},
            {"key": "company_vehicle_name", "label": "Vehicle Name (Company)"},
            {"key": "company_vehicle_number", "label": "Vehicle Number (Company)"},
            {"key": "reach_time", "label": "Reach Time"},
            {"key": "drop_time", "label": "Drop Time"},
            {"key": "company_kms_travelled", "label": "KMs Travelled (Company)"},
            {"key": "company_toll", "label": "Toll (Company)"},
            {"key": "company_diesel", "label": "Diesel (Company)"},
            {"key": "incentive_amount", "label": "Incentive Amount"},
            {"key": "incentive_reason", "label": "Incentive Reason"},
            {"key": "incentive_added_by", "label": "Incentive Added By"},
        ]
        
        col_map = {col["key"]: col["label"] for col in available_columns}
        
        # Write headers
        headers = [col_map.get(col, col) for col in selected_columns]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data rows
        for order in orders:
            row = []
            for col in selected_columns:
                value = order.get(col, "")
                if isinstance(value, (int, float)):
                    row.append(value)
                else:
                    row.append(str(value) if value else "")
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="EXPORT",
            resource_type="REPORT",
            new_data={"report_type": "custom_columns", "format": "excel", "columns": selected_columns}
        )
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=custom_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting custom report: {str(e)}")

@api_router.post("/reports/custom-columns/export/pdf")
async def export_custom_columns_pdf(
    report_config: dict,
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Export custom column report to PDF"""
    try:
        start_date_str = report_config.get("start_date")
        end_date_str = report_config.get("end_date")
        selected_columns = report_config.get("columns", [])
        order_type_filter = report_config.get("order_type", "all")
        
        if not selected_columns:
            raise HTTPException(status_code=400, detail="At least one column must be selected")
        
        # Parse dates
        start = datetime.fromisoformat(start_date_str.replace('Z', '+00:00')) if start_date_str else datetime(2020, 1, 1, tzinfo=timezone.utc)
        end = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')) if end_date_str else datetime.now(timezone.utc)
        
        # Build query
        query = {
            "date_time": {
                "$gte": start.isoformat(),
                "$lte": end.isoformat()
            }
        }
        
        if order_type_filter != "all":
            query["order_type"] = order_type_filter
        
        # Build projection for selected columns
        projection = {"_id": 0, "id": 1}
        for col in selected_columns:
            projection[col] = 1
        
        # Fetch orders with only selected columns (limit to 1000 for PDF)
        orders = await db.crane_orders.find(query, projection).limit(1000).to_list(1000)
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=1
        )
        elements.append(Paragraph("Custom Column Report", title_style))
        elements.append(Spacer(1, 12))
        
        # Date range
        date_text = f"Period: {start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Get column labels
        available_columns = [
            {"key": "unique_id", "label": "Order ID"},
            {"key": "date_time", "label": "Date & Time"},
            {"key": "customer_name", "label": "Customer"},
            {"key": "phone", "label": "Phone"},
            {"key": "order_type", "label": "Type"},
            {"key": "created_by", "label": "Created By"},
            {"key": "cash_trip_from", "label": "From (Cash)"},
            {"key": "cash_trip_to", "label": "To (Cash)"},
            {"key": "cash_driver_name", "label": "Driver (Cash)"},
            {"key": "cash_service_type", "label": "Service (Cash)"},
            {"key": "amount_received", "label": "Amount"},
            {"key": "company_name", "label": "Company"},
            {"key": "company_service_type", "label": "Service (Co)"},
            {"key": "company_driver_name", "label": "Driver (Co)"},
        ]
        
        col_map = {col["key"]: col["label"] for col in available_columns}
        
        # Create table data - limit columns to fit PDF width
        max_cols = min(len(selected_columns), 6)  # Limit to 6 columns for better PDF layout
        display_columns = selected_columns[:max_cols]
        
        table_data = [[col_map.get(col, col) for col in display_columns]]
        
        for order in orders[:100]:  # Limit to 100 rows for PDF
            row = []
            for col in display_columns:
                value = order.get(col, "")
                if isinstance(value, (int, float)):
                    row.append(str(value))
                else:
                    val_str = str(value) if value else "-"
                    # Truncate long values
                    row.append(val_str[:30] + "..." if len(val_str) > 30 else val_str)
            table_data.append(row)
        
        # Create table
        col_widths = [A4[0] / max_cols * 0.9] * max_cols
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        if len(orders) > 100:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Showing first 100 of {len(orders)} records", styles['Italic']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="EXPORT",
            resource_type="REPORT",
            new_data={"report_type": "custom_columns", "format": "pdf", "columns": selected_columns}
        )
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=custom_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.pdf"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting custom report PDF: {str(e)}")

@api_router.get("/reports/available-columns")
async def get_available_columns(
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN]))
):
    """Get list of all available columns for custom reports"""
    columns = [
        {"key": "unique_id", "label": "Order ID", "category": "Basic"},
        {"key": "date_time", "label": "Date & Time", "category": "Basic"},
        {"key": "customer_name", "label": "Customer Name", "category": "Basic"},
        {"key": "phone", "label": "Phone", "category": "Basic"},
        {"key": "order_type", "label": "Order Type", "category": "Basic"},
        {"key": "created_by", "label": "Created By", "category": "Basic"},
        
        # Cash Order Fields
        {"key": "cash_trip_from", "label": "Trip From (Cash)", "category": "Cash"},
        {"key": "cash_trip_to", "label": "Trip To (Cash)", "category": "Cash"},
        {"key": "cash_driver_name", "label": "Driver (Cash)", "category": "Cash"},
        {"key": "cash_towing_vehicle", "label": "Towing Vehicle (Cash)", "category": "Cash"},
        {"key": "cash_service_type", "label": "Service Type (Cash)", "category": "Cash"},
        {"key": "cash_vehicle_name", "label": "Vehicle Name (Cash)", "category": "Cash"},
        {"key": "cash_vehicle_number", "label": "Vehicle Number (Cash)", "category": "Cash"},
        {"key": "amount_received", "label": "Amount Received", "category": "Cash"},
        {"key": "advance_amount", "label": "Advance Amount", "category": "Cash"},
        {"key": "cash_kms_travelled", "label": "KMs Travelled (Cash)", "category": "Cash"},
        {"key": "cash_toll", "label": "Toll (Cash)", "category": "Cash"},
        {"key": "cash_diesel", "label": "Diesel (Cash)", "category": "Cash"},
        {"key": "cash_diesel_refill_location", "label": "Diesel Refill Location", "category": "Cash"},
        {"key": "care_off", "label": "Care Off", "category": "Cash"},
        {"key": "care_off_amount", "label": "Care Off Amount", "category": "Cash"},
        
        # Company Order Fields
        {"key": "name_of_firm", "label": "Name of Firm", "category": "Company"},
        {"key": "company_name", "label": "Company Name", "category": "Company"},
        {"key": "case_id_file_number", "label": "Case ID/File Number", "category": "Company"},
        {"key": "company_trip_from", "label": "Trip From (Company)", "category": "Company"},
        {"key": "company_trip_to", "label": "Trip To (Company)", "category": "Company"},
        {"key": "company_driver_name", "label": "Driver (Company)", "category": "Company"},
        {"key": "company_towing_vehicle", "label": "Towing Vehicle (Company)", "category": "Company"},
        {"key": "company_service_type", "label": "Service Type (Company)", "category": "Company"},
        {"key": "company_vehicle_name", "label": "Vehicle Name (Company)", "category": "Company"},
        {"key": "company_vehicle_number", "label": "Vehicle Number (Company)", "category": "Company"},
        {"key": "reach_time", "label": "Reach Time", "category": "Company"},
        {"key": "drop_time", "label": "Drop Time", "category": "Company"},
        {"key": "company_kms_travelled", "label": "KMs Travelled (Company)", "category": "Company"},
        {"key": "company_toll", "label": "Toll (Company)", "category": "Company"},
        {"key": "company_diesel", "label": "Diesel (Company)", "category": "Company"},
        
        # Incentives
        {"key": "incentive_amount", "label": "Incentive Amount", "category": "Incentives"},
        {"key": "incentive_reason", "label": "Incentive Reason", "category": "Incentives"},
        {"key": "incentive_added_by", "label": "Incentive Added By", "category": "Incentives"},
    ]
    
    return {"columns": columns}

# Data import endpoint
@api_router.post("/import/excel")
async def import_excel_data(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ADMIN])),
):
    """Import orders from Excel file (Admin and Super Admin only)"""
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
        
        # Read the Excel file
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Log headers for debugging
        logging.info(f"Excel file headers: {headers}")
        
        # Create case-insensitive header mapping
        header_map = {}
        for i, header in enumerate(headers):
            if header:
                header_map[str(header).strip().lower()] = i
        
        imported_count = 0
        failed_count = 0
        errors = []
        
        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create a dictionary from row data
                row_data = dict(zip(headers, row))
                
                # Skip empty rows
                if not any(row_data.values()):
                    continue
                
                # Helper function to safely convert values
                def safe_str(val, default=""):
                    if val is None or (isinstance(val, str) and val.strip().lower() in ['nan', 'nat', 'none', '']):
                        return default
                    return str(val).strip()
                
                def safe_float(val, default=0.0):
                    if val is None or val == '':
                        return default
                    try:
                        # Handle string representations of numbers with currency symbols
                        if isinstance(val, str):
                            val = val.replace('₹', '').replace(',', '').strip()
                        return float(val)
                    except (ValueError, TypeError):
                        return default
                
                # Base order data - required fields
                order_data = {
                    "id": str(uuid.uuid4()),
                    "added_time": datetime.now(timezone.utc).isoformat(),
                    "unique_id": safe_str(row_data.get("Order ID") or row_data.get("unique_id"), f"IMP-{uuid.uuid4().hex[:8]}"),
                    "date_time": datetime.now(timezone.utc).isoformat(),  # Will be stored as ISO string
                    "customer_name": safe_str(row_data.get("Customer Name") or row_data.get("customer_name"), "Unknown"),
                    "phone": safe_str(row_data.get("Phone") or row_data.get("phone"), ""),
                    "order_type": safe_str(row_data.get("Order Type") or row_data.get("order_type"), "cash").lower(),
                    "created_by": "system_import"
                }
                
                # Add cash-specific fields
                if order_data["order_type"] == "cash":
                    order_data.update({
                        "cash_trip_from": safe_str(row_data.get("Trip From") or row_data.get("cash_trip_from")),
                        "cash_trip_to": safe_str(row_data.get("Trip To") or row_data.get("cash_trip_to")),
                        "cash_driver_name": safe_str(row_data.get("Driver") or row_data.get("cash_driver_name")),
                        "cash_towing_vehicle": safe_str(row_data.get("Towing Vehicle") or row_data.get("cash_towing_vehicle")),
                        "cash_service_type": safe_str(row_data.get("Service Type") or row_data.get("cash_service_type")),
                        "cash_vehicle_name": safe_str(row_data.get("Vehicle Name") or row_data.get("cash_vehicle_name")),
                        "cash_vehicle_number": safe_str(row_data.get("Vehicle Number") or row_data.get("cash_vehicle_number")),
                        "amount_received": safe_float(row_data.get("Amount Received") or row_data.get("amount_received")),
                        "advance_amount": safe_float(row_data.get("Advance Amount") or row_data.get("advance_amount")),
                        "cash_kms_travelled": safe_float(row_data.get("KMs Travelled") or row_data.get("cash_kms_travelled")),
                        "cash_toll": safe_float(row_data.get("Toll") or row_data.get("cash_toll")),
                        "cash_diesel": safe_float(row_data.get("Diesel") or row_data.get("cash_diesel")),
                        "cash_diesel_refill_location": safe_str(row_data.get("Diesel Location") or row_data.get("cash_diesel_refill_location")),
                    })
                else:  # company order
                    order_data.update({
                        "company_name": safe_str(row_data.get("Company") or row_data.get("company_name")),
                        "case_id_file_number": safe_str(row_data.get("Case ID") or row_data.get("case_id_file_number")),
                        "company_trip_from": safe_str(row_data.get("Trip From") or row_data.get("company_trip_from")),
                        "company_trip_to": safe_str(row_data.get("Trip To") or row_data.get("company_trip_to")),
                        "company_driver_name": safe_str(row_data.get("Driver") or row_data.get("company_driver_name")),
                        "company_towing_vehicle": safe_str(row_data.get("Towing Vehicle") or row_data.get("company_towing_vehicle")),
                        "company_service_type": safe_str(row_data.get("Service Type") or row_data.get("company_service_type")),
                        "company_vehicle_name": safe_str(row_data.get("Vehicle Name") or row_data.get("company_vehicle_name")),
                        "company_vehicle_number": safe_str(row_data.get("Vehicle Number") or row_data.get("company_vehicle_number")),
                        "company_kms_travelled": safe_float(row_data.get("KMs Travelled") or row_data.get("company_kms_travelled")),
                        "company_toll": safe_float(row_data.get("Toll") or row_data.get("company_toll")),
                        "company_diesel": safe_float(row_data.get("Diesel") or row_data.get("company_diesel")),
                        "name_of_firm": safe_str(row_data.get("Firm") or row_data.get("name_of_firm"), "Kawale Cranes"),
                    })
                
                # Insert directly to database without Pydantic validation
                # This allows more flexible import of data
                await db.crane_orders.insert_one(order_data)
                imported_count += 1
                
            except Exception as row_error:
                failed_count += 1
                error_msg = f"Row {row_idx}: {str(row_error)}"
                errors.append(error_msg)
                # Log detailed error for debugging
                logging.error(f"Import error - {error_msg}")
                continue
        
        # Save import history
        import_history = {
            "id": str(uuid.uuid4()),
            "filename": file.filename,
            "imported_at": datetime.now(timezone.utc).isoformat(),
            "imported_by": current_user["email"],
            "total_records": imported_count + failed_count,
            "successful": imported_count,
            "failed": failed_count,
            "status": "completed" if failed_count == 0 else "completed_with_errors"
        }
        await db.import_history.insert_one(import_history)
        
        # Log audit
        await log_audit(
            user_id=current_user["id"],
            user_email=current_user["email"],
            action="IMPORT",
            resource_type="ORDER",
            new_data={
                "filename": file.filename,
                "imported": imported_count,
                "failed": failed_count
            }
        )
        
        result_message = f"Import completed! {imported_count} records imported successfully"
        if failed_count > 0:
            result_message += f", {failed_count} records failed"
        
        return {
            "message": result_message,
            "imported": imported_count,
            "failed": failed_count,
            "total_rows": imported_count + failed_count,
            "errors": errors if len(errors) <= 50 else errors[:50] + [f"... and {len(errors) - 50} more errors"]
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error importing Excel file: {str(e)}")

# Service rates calculation functions
async def initialize_service_rates():
    """Initialize service rates from SK_Rates data"""
    try:
        # Check if rates already exist
        existing_rates = await db.service_rates.count_documents({})
        if existing_rates > 0:
            return  # Rates already initialized
        
        # SK_Rates data based on the provided Excel file
        rates_data = [
            # Kawale Cranes - Europ Assistance
            {"name_of_firm": "Kawale Cranes", "company_name": "Europ Assistance", "service_type": "2 Wheeler Towing", "base_rate": 1200, "rate_per_km_beyond": 12},
            {"name_of_firm": "Kawale Cranes", "company_name": "Europ Assistance", "service_type": "Under-lift", "base_rate": 1500, "rate_per_km_beyond": 15},
            {"name_of_firm": "Kawale Cranes", "company_name": "Europ Assistance", "service_type": "FBT", "base_rate": 1900, "rate_per_km_beyond": 21},
            
            # Vidharbha Towing - Europ Assistance  
            {"name_of_firm": "Vidharbha Towing", "company_name": "Europ Assistance", "service_type": "2 Wheeler Towing", "base_rate": 1200, "rate_per_km_beyond": 12},
            {"name_of_firm": "Vidharbha Towing", "company_name": "Europ Assistance", "service_type": "Under-lift", "base_rate": 1500, "rate_per_km_beyond": 15},
            {"name_of_firm": "Vidharbha Towing", "company_name": "Europ Assistance", "service_type": "FBT", "base_rate": 1900, "rate_per_km_beyond": 21},
            
            # Vira Towing - Europ Assistance
            {"name_of_firm": "Vira Towing", "company_name": "Europ Assistance", "service_type": "2 Wheeler Towing", "base_rate": 1200, "rate_per_km_beyond": 12},
            {"name_of_firm": "Vira Towing", "company_name": "Europ Assistance", "service_type": "Under-lift", "base_rate": 1500, "rate_per_km_beyond": 15},
            {"name_of_firm": "Vira Towing", "company_name": "Europ Assistance", "service_type": "FBT", "base_rate": 1900, "rate_per_km_beyond": 21},
            
            # Sarang Cranes - Europ Assistance
            {"name_of_firm": "Sarang Cranes", "company_name": "Europ Assistance", "service_type": "2 Wheeler Towing", "base_rate": 1200, "rate_per_km_beyond": 12},
            {"name_of_firm": "Sarang Cranes", "company_name": "Europ Assistance", "service_type": "Under-lift", "base_rate": 1500, "rate_per_km_beyond": 15},
            {"name_of_firm": "Sarang Cranes", "company_name": "Europ Assistance", "service_type": "FBT", "base_rate": 1900, "rate_per_km_beyond": 21},
            
            # Kawale Cranes - Mondial
            {"name_of_firm": "Kawale Cranes", "company_name": "Mondial", "service_type": "2 Wheeler Towing", "base_rate": 1200, "rate_per_km_beyond": 13},
            {"name_of_firm": "Kawale Cranes", "company_name": "Mondial", "service_type": "Under-lift", "base_rate": 1400, "rate_per_km_beyond": 15},
            
            # Vidharbha Towing - TVS
            {"name_of_firm": "Vidharbha Towing", "company_name": "TVS", "service_type": "Under-lift", "base_rate": 1700, "rate_per_km_beyond": 17},
            {"name_of_firm": "Vidharbha Towing", "company_name": "TVS", "service_type": "FBT", "base_rate": 2000, "rate_per_km_beyond": 20},
            
            # Vidharbha Towing - Mondial
            {"name_of_firm": "Vidharbha Towing", "company_name": "Mondial", "service_type": "FBT", "base_rate": 1700, "rate_per_km_beyond": 20},
        ]
        
        # Convert to ServiceRate objects and insert
        service_rates = []
        for rate_data in rates_data:
            service_rate = ServiceRate(**rate_data)
            service_rates.append(prepare_for_mongo(service_rate.model_dump()))
        
        await db.service_rates.insert_many(service_rates)
        logging.info(f"Initialized {len(service_rates)} service rates")
        
    except Exception as e:
        logging.error(f"Error initializing service rates: {str(e)}")

async def calculate_order_financials(order: dict) -> CalculatedOrderFinancials:
    """Calculate financial details for a company order based on rates"""
    try:
        # Only calculate for company orders
        if order.get("order_type") != "company":
            return CalculatedOrderFinancials()
        
        # Extract required fields
        name_of_firm = order.get("name_of_firm", "")
        company_name = order.get("company_name", "")
        service_type = order.get("company_service_type", "")
        kms_travelled = order.get("company_kms_travelled", 0) or 0
        incentive_amount = order.get("incentive_amount", 0) or 0
        
        if not all([name_of_firm, company_name, service_type]):
            return CalculatedOrderFinancials(
                incentive_amount=incentive_amount,
                total_revenue=incentive_amount,
                calculation_details="Missing firm, company, or service type information"
            )
        
        # Find matching rate
        rate_query = {
            "name_of_firm": name_of_firm,
            "company_name": company_name,
            "service_type": service_type
        }
        
        rate_doc = await db.service_rates.find_one(rate_query, {"_id": 0})
        
        if not rate_doc:
            return CalculatedOrderFinancials(
                incentive_amount=incentive_amount,
                total_revenue=incentive_amount,
                calculation_details=f"No rate found for {name_of_firm} - {company_name} - {service_type}"
            )
        
        # Calculate base revenue
        base_rate = rate_doc.get("base_rate", 0)
        base_distance = rate_doc.get("base_distance_km", 40)
        rate_per_km_beyond = rate_doc.get("rate_per_km_beyond", 0)
        
        if kms_travelled <= base_distance:
            base_revenue = base_rate
            calculation_details = f"Base rate: ₹{base_rate} for {kms_travelled}km (≤{base_distance}km)"
        else:
            excess_km = kms_travelled - base_distance
            excess_charges = excess_km * rate_per_km_beyond
            base_revenue = base_rate + excess_charges
            calculation_details = f"Base: ₹{base_rate} + Extra {excess_km}km × ₹{rate_per_km_beyond} = ₹{base_revenue}"
        
        total_revenue = base_revenue + incentive_amount
        
        if incentive_amount > 0:
            calculation_details += f" + Incentive: ₹{incentive_amount}"
        
        return CalculatedOrderFinancials(
            base_revenue=base_revenue,
            incentive_amount=incentive_amount,
            total_revenue=total_revenue,
            calculation_details=calculation_details
        )
        
    except Exception as e:
        logging.error(f"Error calculating order financials: {str(e)}")
        return CalculatedOrderFinancials(
            incentive_amount=order.get("incentive_amount", 0) or 0,
            total_revenue=order.get("incentive_amount", 0) or 0,
            calculation_details=f"Calculation error: {str(e)}"
        )

# Initialize default super admin user
async def create_default_super_admin():
    """Create default super admin user if none exists"""
    existing_admin = await db.users.find_one({"role": "super_admin"})
    if not existing_admin:
        default_admin = User(
            email="admin@kawalecranes.com",
            full_name="Super Administrator",
            role=UserRole.SUPER_ADMIN
        )
        
        admin_doc = prepare_for_mongo(default_admin.model_dump())
        admin_doc["hashed_password"] = get_password_hash("admin123")
        
        await db.users.insert_one(admin_doc)
        print("Default super admin created: admin@kawalecranes.com / admin123")

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    await create_default_super_admin()
    await initialize_service_rates()
    
    # Seed database with Excel data if empty (run in background)
    try:
        from seed_database import seed_database_if_empty
        # Run seeding in a separate thread to avoid blocking startup
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor() as executor:
            await loop.run_in_executor(executor, seed_database_if_empty)
    except Exception as e:
        logging.error(f"Error during database seeding: {str(e)}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
